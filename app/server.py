import csv
import base64
import hashlib
import hmac
import json
import os
import re
import secrets
import sqlite3
import unicodedata
import urllib.parse
from datetime import date, datetime, timedelta
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent.parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR / "data"))
DB_PATH = DATA_DIR / "financeiro.db"
TEMP_DIR = DATA_DIR / "tmp"
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)
SESSION_HOURS = int(os.environ.get("SESSION_HOURS", "12"))
SECRET_KEY = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")
SUPERADMIN_USERNAME = "krisrosa"
SUPERADMIN_PASSWORD = os.environ.get("SUPERADMIN_PASSWORD") or "admin123"

DEFAULT_CATEGORIES = [
    "Moradia",
    "Alimentação",
    "Transporte",
    "Saúde",
    "Educação",
    "Cartão de Crédito",
    "Consórcio",
    "Lazer",
    "Investimentos",
    "Impostos",
    "Outros",
]

IMPORT_SHEETS = {
    "DESPESAS": {
        "headers": ["data_lancamento", "descricao", "categoria", "valor", "data_vencimento", "status", "forma_pagamento", "observacao"],
        "required_headers": ["data_lancamento", "descricao", "categoria", "valor", "data_vencimento", "status", "observacao"],
        "aliases": {"data": "data_lancamento", "vencimento": "data_vencimento"},
        "model_headers": ["Data", "Descrição", "Categoria", "Valor", "Vencimento", "Status", "Observação"],
        "sample": ["2026-06-10", "Mercado do mes", "Alimentação", "250.00", "2026-06-15", "Pendente", ""],
    },
    "RECEITAS": {
        "headers": ["data_recebimento", "descricao", "categoria", "valor", "status", "observacao", "recorrente", "quantidade_parcelas", "data_primeiro_recebimento"],
        "required_headers": ["data_recebimento", "descricao", "categoria", "valor", "observacao"],
        "aliases": {"data": "data_recebimento", "quantidade": "quantidade_parcelas", "parcelas": "quantidade_parcelas"},
        "model_headers": ["Data", "Descrição", "Categoria", "Valor", "Observação", "Recorrente", "Quantidade Parcelas", "Data Primeiro Recebimento"],
        "sample": ["2026-06-05", "Salario", "Outros", "3500.00", "", "Nao", "", ""],
    },
    "METAS": {
        "headers": ["nome", "descricao", "valor_objetivo", "valor_atual", "data_prevista", "status"],
        "required_headers": ["nome", "descricao", "valor_objetivo", "valor_atual", "data_prevista"],
        "aliases": {},
        "model_headers": ["Nome", "Descrição", "Valor Objetivo", "Valor Atual", "Data Prevista"],
        "sample": ["Reserva de Emergencia", "Meta inicial", "10000.00", "500.00", "2026-12-31"],
    },
    "PARCELADAS": {
        "headers": ["descricao", "categoria", "valor_parcela", "quantidade_parcelas", "data_primeira_parcela", "status", "forma_pagamento", "observacao"],
        "required_headers": ["descricao", "categoria", "valor_parcela", "quantidade_parcelas", "data_primeira_parcela"],
        "aliases": {"parcelas": "quantidade_parcelas", "valor_total": "valor_parcela"},
        "model_headers": ["Descrição", "Categoria", "Valor Parcela", "Quantidade Parcelas", "Data Primeira Parcela"],
        "sample": ["Cartão Nubank", "Cartão de Crédito", "11.50", "3", "2026-07-10"],
    },
}

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class PostgresCursor:
    def __init__(self, cursor):
        self.cursor = cursor
        self.lastrowid = None

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()

    @property
    def rowcount(self):
        return self.cursor.rowcount


class PostgresConnection:
    def __init__(self):
        try:
            import psycopg2
            import psycopg2.extras
        except ImportError as exc:
            raise RuntimeError("Instale psycopg2-binary ou use a imagem Docker atualizada para conectar ao PostgreSQL") from exc
        self.conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        if exc_type:
            self.conn.rollback()
        else:
            self.conn.commit()
        self.conn.close()

    def execute(self, sql, params=()):
        sql = sql.replace("?", "%s")
        returning = False
        upper_sql = sql.upper()
        table_match = re.match(r"\s*INSERT\s+INTO\s+([a-z_]+)", sql, re.IGNORECASE)
        id_tables = {"users", "categories", "expenses", "incomes", "goals"}
        insert_table = table_match.group(1).lower() if table_match else ""
        if insert_table in id_tables and " RETURNING " not in upper_sql and " ON CONFLICT " not in upper_sql:
            sql = sql.rstrip().rstrip(";") + " RETURNING id"
            returning = True
        cursor = self.conn.cursor()
        cursor.execute(sql, params)
        wrapped = PostgresCursor(cursor)
        if returning:
            try:
                row = cursor.fetchone()
                wrapped.lastrowid = row["id"] if row and "id" in row else None
            except Exception:
                wrapped.lastrowid = None
        return wrapped

    def executescript(self, script):
        cursor = self.conn.cursor()
        cursor.execute(script)
        return PostgresCursor(cursor)


def connect():
    if USE_POSTGRES:
        return PostgresConnection()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return salt, digest.hex()


def verify_password(password, salt, password_hash):
    _, candidate = hash_password(password, salt)
    return hmac.compare_digest(candidate, password_hash)


def integrity_errors():
    errors = [sqlite3.IntegrityError]
    if USE_POSTGRES:
        try:
            import psycopg2
            errors.append(psycopg2.IntegrityError)
        except ImportError:
            pass
    return tuple(errors)


def insert_default_category(conn, category):
    if USE_POSTGRES:
        conn.execute("INSERT INTO categories (name, active) VALUES (?, 1) ON CONFLICT (name) DO NOTHING", (category,))
    else:
        conn.execute("INSERT OR IGNORE INTO categories (name, active) VALUES (?, 1)", (category,))


SQLITE_SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL UNIQUE,
    full_name TEXT NOT NULL DEFAULT '',
    password_hash TEXT NOT NULL,
    salt TEXT NOT NULL,
    profile TEXT NOT NULL DEFAULT 'usuario',
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS sessions (
    token TEXT PRIMARY KEY,
    user_id INTEGER NOT NULL,
    expires_at TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS expenses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    description TEXT NOT NULL,
    category TEXT NOT NULL,
    amount REAL NOT NULL,
    due_date TEXT NOT NULL,
    status TEXT NOT NULL CHECK(status IN ('Pendente', 'Pago', 'Vencido')),
    payment_method TEXT,
    notes TEXT,
    payment_date TEXT,
    user_id INTEGER,
    installment_group TEXT,
    installment_number INTEGER,
    installment_total INTEGER,
    cancelled INTEGER NOT NULL DEFAULT 0,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS incomes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    description TEXT NOT NULL,
    category TEXT NOT NULL,
    amount REAL NOT NULL,
    receipt_date TEXT NOT NULL,
    status TEXT NOT NULL CHECK(status IN ('Recebido', 'Pendente')),
    notes TEXT,
    user_id INTEGER,
    recurring_group TEXT,
    recurring_number INTEGER,
    recurring_total INTEGER,
    cancelled INTEGER NOT NULL DEFAULT 0,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS goals (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    name TEXT NOT NULL,
    description TEXT,
    target_amount REAL NOT NULL,
    current_amount REAL NOT NULL DEFAULT 0,
    target_date TEXT,
    status TEXT NOT NULL DEFAULT 'Ativa',
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS maintenance_settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS maintenance_cleanup_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_at TEXT NOT NULL,
    mode TEXT NOT NULL,
    data_type TEXT NOT NULL,
    retention_days INTEGER NOT NULL,
    deleted_count INTEGER NOT NULL,
    details TEXT
);

CREATE TABLE IF NOT EXISTS import_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER,
    filename TEXT,
    imported_count INTEGER NOT NULL DEFAULT 0,
    skipped_count INTEGER NOT NULL DEFAULT 0,
    errors_count INTEGER NOT NULL DEFAULT 0,
    is_test INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users(id)
);
"""


POSTGRES_SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id SERIAL PRIMARY KEY,
    username TEXT NOT NULL UNIQUE,
    full_name TEXT NOT NULL DEFAULT '',
    password_hash TEXT NOT NULL,
    salt TEXT NOT NULL,
    profile TEXT NOT NULL DEFAULT 'usuario',
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS sessions (
    token TEXT PRIMARY KEY,
    user_id INTEGER NOT NULL REFERENCES users(id),
    expires_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS categories (
    id SERIAL PRIMARY KEY,
    name TEXT NOT NULL UNIQUE,
    active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS expenses (
    id SERIAL PRIMARY KEY,
    description TEXT NOT NULL,
    category TEXT NOT NULL,
    amount DOUBLE PRECISION NOT NULL,
    due_date TEXT NOT NULL,
    status TEXT NOT NULL CHECK(status IN ('Pendente', 'Pago', 'Vencido')),
    payment_method TEXT,
    notes TEXT,
    payment_date TEXT,
    user_id INTEGER REFERENCES users(id),
    installment_group TEXT,
    installment_number INTEGER,
    installment_total INTEGER,
    cancelled INTEGER NOT NULL DEFAULT 0,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS incomes (
    id SERIAL PRIMARY KEY,
    description TEXT NOT NULL,
    category TEXT NOT NULL,
    amount DOUBLE PRECISION NOT NULL,
    receipt_date TEXT NOT NULL,
    status TEXT NOT NULL CHECK(status IN ('Recebido', 'Pendente')),
    notes TEXT,
    user_id INTEGER REFERENCES users(id),
    recurring_group TEXT,
    recurring_number INTEGER,
    recurring_total INTEGER,
    cancelled INTEGER NOT NULL DEFAULT 0,
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS goals (
    id SERIAL PRIMARY KEY,
    user_id INTEGER REFERENCES users(id),
    name TEXT NOT NULL,
    description TEXT,
    target_amount DOUBLE PRECISION NOT NULL,
    current_amount DOUBLE PRECISION NOT NULL DEFAULT 0,
    target_date TEXT,
    status TEXT NOT NULL DEFAULT 'Ativa',
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS maintenance_settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS maintenance_cleanup_logs (
    id SERIAL PRIMARY KEY,
    run_at TEXT NOT NULL,
    mode TEXT NOT NULL,
    data_type TEXT NOT NULL,
    retention_days INTEGER NOT NULL,
    deleted_count INTEGER NOT NULL,
    details TEXT
);

CREATE TABLE IF NOT EXISTS import_history (
    id SERIAL PRIMARY KEY,
    user_id INTEGER REFERENCES users(id),
    filename TEXT,
    imported_count INTEGER NOT NULL DEFAULT 0,
    skipped_count INTEGER NOT NULL DEFAULT 0,
    errors_count INTEGER NOT NULL DEFAULT 0,
    is_test INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL
);
"""


def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        if USE_POSTGRES:
            print("Conectado ao PostgreSQL")
        conn.executescript(POSTGRES_SCHEMA if USE_POSTGRES else SQLITE_SCHEMA)
        ensure_user_columns(conn)
        ensure_category_columns(conn)
        ensure_financial_columns(conn)
        ensure_goal_table(conn)
        ensure_maintenance_tables(conn)
        for category in DEFAULT_CATEGORIES:
            insert_default_category(conn, category)

        user = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
        if not user:
            salt, password_hash = hash_password("admin123")
            conn.execute(
                "INSERT INTO users (username, full_name, password_hash, salt, profile, active, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                ("admin", "Administrador", password_hash, salt, "admin", 1, now()),
            )
        else:
            conn.execute("UPDATE users SET profile = 'admin', active = 1 WHERE username = 'admin'")

        superadmin = conn.execute("SELECT id FROM users WHERE username = ?", (SUPERADMIN_USERNAME,)).fetchone()
        if not superadmin:
            salt, password_hash = hash_password(SUPERADMIN_PASSWORD)
            conn.execute(
                "INSERT INTO users (username, full_name, password_hash, salt, profile, active, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (SUPERADMIN_USERNAME, "Kris Rosa", password_hash, salt, "superadmin", 1, now()),
            )
        else:
            conn.execute(
                "UPDATE users SET profile = 'superadmin', active = 1 WHERE username = ?",
                (SUPERADMIN_USERNAME,),
            )

        admin_id = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()["id"]
        conn.execute("UPDATE expenses SET user_id = ? WHERE user_id IS NULL", (admin_id,))
        conn.execute("UPDATE incomes SET user_id = ? WHERE user_id IS NULL", (admin_id,))
        run_daily_cleanup(conn)
        print("Migração concluída")


def table_columns(conn, table):
    if USE_POSTGRES:
        rows = conn.execute(
            """
            SELECT column_name AS name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = ?
            """,
            (table,),
        ).fetchall()
    else:
        rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return {row["name"] for row in rows}


def ensure_user_columns(conn):
    columns = table_columns(conn, "users")
    if "full_name" not in columns:
        conn.execute("ALTER TABLE users ADD COLUMN full_name TEXT NOT NULL DEFAULT ''")
    if "profile" not in columns:
        conn.execute("ALTER TABLE users ADD COLUMN profile TEXT NOT NULL DEFAULT 'usuario'")
    if "active" not in columns:
        conn.execute("ALTER TABLE users ADD COLUMN active INTEGER NOT NULL DEFAULT 1")


def ensure_category_columns(conn):
    columns = table_columns(conn, "categories")
    if "active" not in columns:
        conn.execute("ALTER TABLE categories ADD COLUMN active INTEGER NOT NULL DEFAULT 1")


def ensure_financial_columns(conn):
    expense_columns = table_columns(conn, "expenses")
    income_columns = table_columns(conn, "incomes")
    if "user_id" not in expense_columns:
        conn.execute("ALTER TABLE expenses ADD COLUMN user_id INTEGER")
    if "user_id" not in income_columns:
        conn.execute("ALTER TABLE incomes ADD COLUMN user_id INTEGER")
    if "active" not in expense_columns:
        conn.execute("ALTER TABLE expenses ADD COLUMN active INTEGER NOT NULL DEFAULT 1")
    if "active" not in income_columns:
        conn.execute("ALTER TABLE incomes ADD COLUMN active INTEGER NOT NULL DEFAULT 1")
    for column, definition in {
        "recurring_group": "TEXT",
        "recurring_number": "INTEGER",
        "recurring_total": "INTEGER",
        "cancelled": "INTEGER NOT NULL DEFAULT 0",
    }.items():
        if column not in income_columns:
            conn.execute(f"ALTER TABLE incomes ADD COLUMN {column} {definition}")
    for column, definition in {
        "installment_group": "TEXT",
        "installment_number": "INTEGER",
        "installment_total": "INTEGER",
        "cancelled": "INTEGER NOT NULL DEFAULT 0",
    }.items():
        if column not in expense_columns:
            conn.execute(f"ALTER TABLE expenses ADD COLUMN {column} {definition}")


def ensure_goal_table(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS goals (
            id SERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id),
            name TEXT NOT NULL,
            description TEXT,
            target_amount DOUBLE PRECISION NOT NULL,
            current_amount DOUBLE PRECISION NOT NULL DEFAULT 0,
            target_date TEXT,
            status TEXT NOT NULL DEFAULT 'Em andamento',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
        if USE_POSTGRES
        else """
        CREATE TABLE IF NOT EXISTS goals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            target_amount REAL NOT NULL,
            current_amount REAL NOT NULL DEFAULT 0,
            target_date TEXT,
            status TEXT NOT NULL DEFAULT 'Em andamento',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        """
    )
    columns = table_columns(conn, "goals")
    for column, definition in {
        "user_id": "INTEGER",
        "description": "TEXT",
        "current_amount": "REAL NOT NULL DEFAULT 0",
        "target_date": "TEXT",
        "status": "TEXT NOT NULL DEFAULT 'Em andamento'",
        "active": "INTEGER NOT NULL DEFAULT 1",
        "created_at": "TEXT",
        "updated_at": "TEXT",
    }.items():
        if column not in columns:
            conn.execute(f"ALTER TABLE goals ADD COLUMN {column} {definition}")


def ensure_maintenance_tables(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS maintenance_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS maintenance_cleanup_logs (
            id SERIAL PRIMARY KEY,
            run_at TEXT NOT NULL,
            mode TEXT NOT NULL,
            data_type TEXT NOT NULL,
            retention_days INTEGER NOT NULL,
            deleted_count INTEGER NOT NULL,
            details TEXT
        )
        """
        if USE_POSTGRES
        else """
        CREATE TABLE IF NOT EXISTS maintenance_cleanup_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_at TEXT NOT NULL,
            mode TEXT NOT NULL,
            data_type TEXT NOT NULL,
            retention_days INTEGER NOT NULL,
            deleted_count INTEGER NOT NULL,
            details TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS import_history (
            id SERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id),
            filename TEXT,
            imported_count INTEGER NOT NULL DEFAULT 0,
            skipped_count INTEGER NOT NULL DEFAULT 0,
            errors_count INTEGER NOT NULL DEFAULT 0,
            is_test INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
        if USE_POSTGRES
        else """
        CREATE TABLE IF NOT EXISTS import_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            filename TEXT,
            imported_count INTEGER NOT NULL DEFAULT 0,
            skipped_count INTEGER NOT NULL DEFAULT 0,
            errors_count INTEGER NOT NULL DEFAULT 0,
            is_test INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
        """
    )
    for table, definitions in {
        "maintenance_cleanup_logs": {
            "run_at": "TEXT",
            "mode": "TEXT NOT NULL DEFAULT 'manual'",
            "data_type": "TEXT NOT NULL DEFAULT 'desconhecido'",
            "retention_days": "INTEGER NOT NULL DEFAULT 0",
            "deleted_count": "INTEGER NOT NULL DEFAULT 0",
            "details": "TEXT",
        },
        "import_history": {
            "user_id": "INTEGER",
            "filename": "TEXT",
            "imported_count": "INTEGER NOT NULL DEFAULT 0",
            "skipped_count": "INTEGER NOT NULL DEFAULT 0",
            "errors_count": "INTEGER NOT NULL DEFAULT 0",
            "is_test": "INTEGER NOT NULL DEFAULT 0",
            "created_at": "TEXT",
        },
    }.items():
        columns = table_columns(conn, table)
        for column, definition in definitions.items():
            if column not in columns:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
    ensure_maintenance_defaults(conn)


DEFAULT_MAINTENANCE_SETTINGS = {
    "cleanup_logs_days": 90,
    "import_history_days": 180,
    "temp_files_days": 30,
    "test_records_days": 30,
}


def table_exists(conn, table):
    if USE_POSTGRES:
        row = conn.execute(
            """
            SELECT table_name AS name
            FROM information_schema.tables
            WHERE table_schema = 'public' AND table_name = ?
            """,
            (table,),
        ).fetchone()
    else:
        row = conn.execute("SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?", (table,)).fetchone()
    return row is not None


def not_deleted_filter(conn, table):
    return " AND deleted = 0" if "deleted" in table_columns(conn, table) else ""


def set_maintenance_setting(conn, key, value):
    conn.execute(
        """
        INSERT INTO maintenance_settings (key, value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at
        """,
        (key, str(value), now()),
    )


def ensure_maintenance_defaults(conn):
    for key, value in DEFAULT_MAINTENANCE_SETTINGS.items():
        existing = conn.execute("SELECT value FROM maintenance_settings WHERE key = ?", (key,)).fetchone()
        if not existing:
            set_maintenance_setting(conn, key, value)


def maintenance_settings(conn):
    ensure_maintenance_defaults(conn)
    rows = conn.execute("SELECT key, value FROM maintenance_settings").fetchall()
    settings = {row["key"]: int(row["value"]) for row in rows if str(row["value"]).isdigit()}
    for key, value in DEFAULT_MAINTENANCE_SETTINGS.items():
        settings.setdefault(key, value)
    return settings


def count_where(conn, table, where, params=()):
    row = conn.execute(f"SELECT COUNT(*) AS total FROM {table} WHERE {where}", params).fetchone()
    return int(row["total"] or 0)


def delete_where(conn, table, where, params=()):
    return conn.execute(f"DELETE FROM {table} WHERE {where}", params).rowcount


def log_cleanup(conn, run_at, mode, data_type, retention_days, deleted_count, details=""):
    conn.execute(
        """
        INSERT INTO maintenance_cleanup_logs
        (run_at, mode, data_type, retention_days, deleted_count, details)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (run_at, mode, data_type, retention_days, int(deleted_count), details),
    )


def cleanup_table_before(conn, run_at, mode, table, data_type, date_column, days, extra_where="", extra_params=(), details=""):
    cutoff = (datetime.now() - timedelta(days=int(days))).replace(microsecond=0).isoformat()
    where = f"{date_column} < ?"
    params = [cutoff]
    if extra_where:
        where += f" AND {extra_where}"
        params.extend(extra_params)
    total = count_where(conn, table, where, params)
    log_cleanup(conn, run_at, mode, data_type, days, total, details or f"Registros anteriores a {cutoff}")
    if total:
        delete_where(conn, table, where, params)
    return {"type": data_type, "deleted": total, "retention_days": int(days)}


def cleanup_expired_sessions(conn, run_at, mode):
    current = now()
    total = count_where(conn, "sessions", "expires_at <= ?", (current,))
    log_cleanup(conn, run_at, mode, "sessions_expiradas", 0, total, f"Sessoes vencidas ate {current}")
    if total:
        delete_where(conn, "sessions", "expires_at <= ?", (current,))
    return {"type": "sessions_expiradas", "deleted": total, "retention_days": 0}


def cleanup_expired_tokens(conn, run_at, mode):
    if not table_exists(conn, "tokens") or "expires_at" not in table_columns(conn, "tokens"):
        log_cleanup(conn, run_at, mode, "tokens_vencidos", 0, 0, "Tabela tokens nao existe")
        return {"type": "tokens_vencidos", "deleted": 0, "retention_days": 0}
    current = now()
    total = count_where(conn, "tokens", "expires_at <= ?", (current,))
    log_cleanup(conn, run_at, mode, "tokens_vencidos", 0, total, f"Tokens vencidos ate {current}")
    if total:
        delete_where(conn, "tokens", "expires_at <= ?", (current,))
    return {"type": "tokens_vencidos", "deleted": total, "retention_days": 0}


def cleanup_temp_files(conn, run_at, mode, days):
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    cutoff = datetime.now() - timedelta(days=int(days))
    files = []
    for path in TEMP_DIR.rglob("*"):
        if path.is_file() and datetime.fromtimestamp(path.stat().st_mtime) < cutoff:
            files.append(path)
    log_cleanup(conn, run_at, mode, "arquivos_temporarios", days, len(files), f"Diretorio: {TEMP_DIR}")
    for path in files:
        try:
            path.unlink()
        except OSError as exc:
            print(f"[manutencao] falha ao remover arquivo temporario {path}: {exc}")
    return {"type": "arquivos_temporarios", "deleted": len(files), "retention_days": int(days)}


def run_cleanup(conn, mode="manual"):
    settings = maintenance_settings(conn)
    run_at = now()
    report = [
        cleanup_expired_sessions(conn, run_at, mode),
        cleanup_expired_tokens(conn, run_at, mode),
        cleanup_table_before(conn, run_at, mode, "maintenance_cleanup_logs", "logs_manutencao", "run_at", settings["cleanup_logs_days"]),
        cleanup_table_before(conn, run_at, mode, "import_history", "registros_teste", "created_at", settings["test_records_days"], "is_test = 1", details="Somente historico de importacao marcado como teste"),
        cleanup_table_before(conn, run_at, mode, "import_history", "historico_importacoes", "created_at", settings["import_history_days"]),
        cleanup_temp_files(conn, run_at, mode, settings["temp_files_days"]),
    ]
    return {"run_at": run_at, "mode": mode, "report": report}


def run_daily_cleanup(conn):
    last = conn.execute(
        "SELECT run_at FROM maintenance_cleanup_logs WHERE mode = ? ORDER BY run_at DESC LIMIT 1",
        ("automatico",),
    ).fetchone()
    if last and str(last["run_at"])[:10] == today_iso():
        return
    result = run_cleanup(conn, "automatico")
    print(f"[manutencao] limpeza automatica concluida em {result['run_at']}")


def now():
    return datetime.now().replace(microsecond=0).isoformat()


def today_iso():
    return date.today().isoformat()


def month_range(year, month):
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)
    return start.isoformat(), end.isoformat()


def add_months(value, months):
    current = datetime.strptime(value, "%Y-%m-%d").date()
    month_index = current.month - 1 + months
    year = current.year + month_index // 12
    month = month_index % 12 + 1
    days = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(current.day, days[month - 1])
    return date(year, month, day).isoformat()


def iter_months(start_date, total):
    for offset in range(total):
        month_start = date(start_date.year, start_date.month, 1)
        yield add_months(month_start.isoformat(), offset)[:7]


def parse_body(handler):
    length = int(handler.headers.get("Content-Length", "0"))
    if length == 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw or "{}")


def row_to_dict(row):
    return dict(row) if row else None


def build_xlsx(sheets):
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=name)
        for row in rows:
            worksheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_import_template():
    return build_xlsx({name: [config.get("model_headers", config["headers"]), config["sample"]] for name, config in IMPORT_SHEETS.items()})


def normalize_header(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def canonical_import_header(sheet_config, value):
    header = normalize_header(value)
    return sheet_config.get("aliases", {}).get(header, header)


def read_xlsx_workbook(payload):
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    except (InvalidFileException, KeyError, OSError, ValueError) as exc:
        print(f"[importacao] erro ao abrir XLSX com openpyxl.load_workbook: {type(exc).__name__}: {exc}")
        print(f"[importacao] tamanho do arquivo recebido: {len(payload)} bytes")
        raise ValueError("Arquivo XLSX invalido ou corrompido. Baixe novamente o modelo oficial e tente importar outra vez.") from exc
    except Exception as exc:
        print(f"[importacao] erro inesperado ao ler XLSX: {type(exc).__name__}: {exc}")
        print(f"[importacao] tamanho do arquivo recebido: {len(payload)} bytes")
        raise ValueError("Nao foi possivel ler a planilha Excel. Verifique se o arquivo esta no formato .xlsx.") from exc

    print(f"[importacao] workbook carregado com openpyxl. Abas encontradas: {workbook.sheetnames}")
    sheets = {}
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            values = []
            for value in row:
                if value is None:
                    values.append("")
                elif isinstance(value, datetime):
                    values.append(value.date().isoformat())
                elif isinstance(value, date):
                    values.append(value.isoformat())
                else:
                    values.append(str(value).strip())
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        sheets[worksheet.title.upper()] = rows
        print(f"[importacao] aba {worksheet.title}: {len(rows)} linhas lidas")
    return sheets


def public_user(row):
    data = row_to_dict(row)
    if not data:
        return None
    return {
        "id": data["id"],
        "username": data["username"],
        "full_name": data.get("full_name", ""),
        "profile": data.get("profile", "usuario"),
        "active": bool(data.get("active", 1)),
        "created_at": data.get("created_at"),
    }


def clean_money(value):
    if isinstance(value, str):
        value = value.strip().replace("R$", "").replace(" ", "")
        if "," in value and "." in value:
            value = value.replace(".", "").replace(",", ".")
        elif "," in value:
            value = value.replace(",", ".")
    try:
        amount = float(value)
    except (TypeError, ValueError):
        raise ValueError("Valor invalido")
    if amount < 0:
        raise ValueError("Valor nao pode ser negativo")
    return round(amount, 2)


def clean_date(value):
    value = str(value or "").strip()
    if not value:
        raise ValueError("Data obrigatoria")
    if re.fullmatch(r"\d+(\.0+)?", value):
        base = date(1899, 12, 30)
        return (base + timedelta(days=int(float(value)))).isoformat()
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise ValueError("Data invalida. Use AAAA-MM-DD")
    return value


def add_months(value, months):
    base = datetime.strptime(value, "%Y-%m-%d").date()
    month = base.month - 1 + months
    year = base.year + month // 12
    month = month % 12 + 1
    max_day = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]
    return date(year, month, min(base.day, max_day)).isoformat()
def clean_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def is_yes(value):
    return str(value or "").strip().lower() in ("sim", "s", "true", "1", "yes")


def base_installment_description(description):
    return re.sub(r"\s*\(\d+\s*/\s*\d+\)\s*$", "", str(description or "").strip())


def require_fields(data, fields):
    missing = [field for field in fields if not str(data.get(field, "")).strip()]
    if missing:
        raise ValueError("Campos obrigatorios: " + ", ".join(missing))


def parse_query(path):
    parsed = urllib.parse.urlparse(path)
    params = urllib.parse.parse_qs(parsed.query)
    return parsed.path, {k: v[0] for k, v in params.items() if v}


def filters_to_sql(params, kind):
    clauses = []
    values = []
    month = params.get("month")
    year = params.get("year")
    status = params.get("status")
    category = params.get("category")
    date_field = "due_date" if kind == "expenses" else "receipt_date"

    if month and year:
        start, end = month_range(int(year), int(month))
        clauses.append(f"{date_field} >= ? AND {date_field} < ?")
        values.extend([start, end])
    elif year:
        clauses.append(f"substr({date_field}, 1, 4) = ?")
        values.append(str(year))

    if status:
        clauses.append("status = ?")
        values.append(status)
    if category:
        clauses.append("category = ?")
        values.append(category)

    where = " WHERE " + " AND ".join(clauses) if clauses else ""
    return where, values


def apply_user_scope(where, values, user, params=None):
    params = params or {}
    selected_user_id = params.get("user_id")
    if user["profile"] == "superadmin" and selected_user_id:
        if where:
            where += " AND user_id = ?"
        else:
            where = " WHERE user_id = ?"
        values.append(selected_user_id)
        return where, values
    if user["profile"] == "superadmin":
        return where, values
    if where:
        where += " AND user_id = ?"
    else:
        where = " WHERE user_id = ?"
    values.append(user["id"])
    return where, values


def owned_update_suffix(user, params=None):
    params = params or {}
    selected_user_id = params.get("user_id")
    if user["profile"] == "superadmin" and selected_user_id:
        return " AND user_id = ?", [selected_user_id]
    if user["profile"] == "superadmin":
        return "", []
    return " AND user_id = ?", [user["id"]]


def normalize_overdue_expenses():
    with connect() as conn:
        conn.execute(
            """
            UPDATE expenses
            SET status = 'Vencido', updated_at = ?
            WHERE status = 'Pendente' AND due_date < ?
            """,
            (now(), today_iso()),
        )


def expense_payload(data, partial=False):
    if not partial:
        require_fields(data, ["description", "category", "amount", "due_date", "status"])
    status = data.get("status", "Pendente")
    if status not in ("Pendente", "Pago", "Vencido"):
        raise ValueError("Status de despesa invalido")
    return {
        "description": str(data.get("description", "")).strip(),
        "category": str(data.get("category", "")).strip(),
        "amount": clean_money(data.get("amount", 0)),
        "due_date": str(data.get("due_date", "")).strip(),
        "status": status,
        "payment_method": str(data.get("payment_method", "")).strip(),
        "notes": str(data.get("notes", "")).strip(),
        "payment_date": str(data.get("payment_date", "")).strip(),
    }


def income_payload(data, partial=False):
    if not partial:
        require_fields(data, ["description", "category", "amount", "receipt_date", "status"])
    status = data.get("status", "Pendente")
    if status not in ("Recebido", "Pendente"):
        raise ValueError("Status de receita invalido")
    return {
        "description": str(data.get("description", "")).strip(),
        "category": str(data.get("category", "")).strip(),
        "amount": clean_money(data.get("amount", 0)),
        "receipt_date": str(data.get("receipt_date", "")).strip(),
        "status": status,
        "notes": str(data.get("notes", "")).strip(),
    }


def goal_payload(data):
    require_fields(data, ["name", "target_amount", "status"])
    status = str(data.get("status", "Ativa")).strip()
    if status not in ("Ativa", "Concluida", "Pausada", "Cancelada"):
        raise ValueError("Status de meta invalido")
    return {
        "name": str(data.get("name", "")).strip(),
        "description": str(data.get("description", "")).strip(),
        "target_amount": clean_money(data.get("target_amount", 0)),
        "current_amount": clean_money(data.get("current_amount", 0)),
        "target_date": str(data.get("target_date", "")).strip(),
        "status": status,
    }


def enrich_goal(goal):
    item = dict(goal)
    target = float(item.get("target_amount") or 0)
    current = float(item.get("current_amount") or 0)
    remaining = max(target - current, 0)
    percent = 100 if target <= 0 else min(round((current / target) * 100, 2), 100)
    item["remaining_amount"] = round(remaining, 2)
    item["percent_complete"] = percent
    item["forecast"] = "Concluida" if remaining <= 0 else (item.get("target_date") or "Sem data")
    return item


def user_payload(data, creating=False, public=False):
    fields = ["full_name", "username"]
    if not public:
        fields.append("profile")
    if creating:
        fields.append("password")
    require_fields(data, fields)
    full_name = str(data.get("full_name", "")).strip()
    username = str(data.get("username", "")).strip()
    profile = "usuario" if public else str(data.get("profile", "usuario")).strip()
    if profile not in ("admin", "usuario"):
        raise ValueError("Perfil invalido")
    active = 1 if str(data.get("active", "true")).lower() in ("1", "true", "sim", "on") else 0
    password = str(data.get("password", ""))
    if creating and len(password) < 6:
        raise ValueError("A senha deve ter pelo menos 6 caracteres")
    confirm_password = str(data.get("confirm_password", password))
    if creating and password != confirm_password:
        raise ValueError("Confirmacao de senha diferente da senha")
    return {"full_name": full_name, "username": username, "profile": profile, "active": active, "password": password}


class FinanceHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def log_message(self, fmt, *args):
        print("%s - %s" % (self.address_string(), fmt % args))

    def send_json(self, data, status=HTTPStatus.OK):
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def send_error_json(self, message, status=HTTPStatus.BAD_REQUEST):
        self.send_json({"error": message}, status)

    def current_user(self):
        token = self.headers.get("Authorization", "").replace("Bearer ", "").strip()
        if not token:
            return None
        with connect() as conn:
            row = conn.execute(
                """
                SELECT users.id, users.username, users.full_name, users.profile, users.active, users.created_at
                FROM sessions
                JOIN users ON users.id = sessions.user_id
                WHERE sessions.token = ? AND sessions.expires_at > ? AND users.active = 1
                """,
                (token, now()),
            ).fetchone()
            return public_user(row)

    def require_user(self):
        user = self.current_user()
        if not user:
            self.send_error_json("Login necessario", HTTPStatus.UNAUTHORIZED)
            return None
        return user

    def require_superadmin(self):
        user = self.require_user()
        if not user:
            return None
        if user["profile"] != "superadmin":
            self.send_error_json("Acesso permitido somente para o SuperAdmin", HTTPStatus.FORBIDDEN)
            return None
        return user

    def require_category_admin(self):
        user = self.require_user()
        if not user:
            return None
        if user["profile"] not in ("superadmin", "admin"):
            self.send_error_json("Acesso permitido somente para administradores", HTTPStatus.FORBIDDEN)
            return None
        return user

    def do_GET(self):
        path, params = parse_query(self.path)
        try:
            if path == "/health":
                return self.send_json({"status": "ok"})
            if path.startswith("/api/"):
                if path == "/api/me":
                    user = self.require_user()
                    return user and self.send_json({"user": user})
                if path == "/api/users":
                    return self.handle_list_users()
                if path == "/api/categories":
                    return self.handle_categories()
                if path == "/api/categories/manage":
                    return self.handle_manage_categories()
                if path == "/api/dashboard":
                    return self.handle_dashboard(params)
                if path == "/api/expenses":
                    return self.handle_list_expenses(params)
                if path == "/api/incomes":
                    return self.handle_list_incomes(params)
                if path == "/api/goals":
                    return self.handle_list_goals(params)
                if path == "/api/report":
                    return self.handle_report(params)
                if path == "/api/report/export":
                    return self.handle_report_export(params)
                if path == "/api/charts":
                    return self.handle_charts(params)
                if path == "/api/cashflow":
                    return self.handle_cashflow(params)
                if path == "/api/import/template":
                    return self.handle_import_template()
                if path == "/api/maintenance":
                    return self.handle_maintenance_status()
                return self.send_error_json("Rota nao encontrada", HTTPStatus.NOT_FOUND)
            return super().do_GET()
        except Exception as exc:
            return self.send_error_json(str(exc), HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_POST(self):
        path, _ = parse_query(self.path)
        try:
            data = parse_body(self)
            if path == "/api/login":
                return self.handle_login(data)
            if path == "/api/register":
                return self.handle_register(data)
            if path == "/api/logout":
                return self.handle_logout()
            if path == "/api/change-password":
                return self.handle_change_password(data)
            if path == "/api/users":
                return self.handle_create_user(data)
            if path == "/api/categories":
                return self.handle_create_category(data)
            if path.startswith("/api/users/") and path.endswith("/password"):
                user_id = int(path.split("/")[3])
                return self.handle_admin_change_password(user_id, data)
            if path.startswith("/api/users/") and path.endswith("/toggle"):
                user_id = int(path.split("/")[3])
                return self.handle_toggle_user(user_id)
            if path == "/api/import/preview":
                return self.handle_import_preview(data)
            if path == "/api/import/commit":
                return self.handle_import_commit(data)
            if path == "/api/maintenance/run":
                return self.handle_run_maintenance()
            if self.require_user() is None:
                return
            if path == "/api/expenses":
                return self.handle_create_expense(data)
            if path == "/api/incomes":
                return self.handle_create_income(data)
            if path == "/api/goals":
                return self.handle_create_goal(data)
            if path == "/api/import/preview":
                return self.handle_import_preview(data)
            if path == "/api/import/commit":
                return self.handle_import_commit(data)
            if path.startswith("/api/expenses/") and path.endswith("/pay"):
                expense_id = int(path.split("/")[3])
                return self.handle_pay_expense(expense_id, data)
            if path.startswith("/api/expenses/") and path.endswith("/cancel-future"):
                expense_id = int(path.split("/")[3])
                return self.handle_cancel_future_installments(expense_id)
            if path.startswith("/api/incomes/") and path.endswith("/receive"):
                income_id = int(path.split("/")[3])
                return self.handle_receive_income(income_id)
            if path.startswith("/api/incomes/") and path.endswith("/cancel-future"):
                income_id = int(path.split("/")[3])
                return self.handle_cancel_future_incomes(income_id)
            if path.startswith("/api/goals/") and path.endswith("/add"):
                goal_id = int(path.split("/")[3])
                return self.handle_goal_amount(goal_id, data, "add")
            if path.startswith("/api/goals/") and path.endswith("/withdraw"):
                goal_id = int(path.split("/")[3])
                return self.handle_goal_amount(goal_id, data, "withdraw")
            return self.send_error_json("Rota nao encontrada", HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            return self.send_error_json(str(exc))
        except Exception as exc:
            return self.send_error_json(str(exc), HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_PUT(self):
        path, _ = parse_query(self.path)
        try:
            if self.require_user() is None:
                return
            data = parse_body(self)
            if path.startswith("/api/users/"):
                return self.handle_update_user(int(path.split("/")[3]), data)
            if path.startswith("/api/categories/"):
                return self.handle_update_category(int(path.split("/")[3]), data)
            if path == "/api/maintenance/settings":
                return self.handle_update_maintenance_settings(data)
            if path.startswith("/api/expenses/"):
                return self.handle_update_expense(int(path.split("/")[3]), data)
            if path.startswith("/api/incomes/"):
                return self.handle_update_income(int(path.split("/")[3]), data)
            if path.startswith("/api/goals/"):
                return self.handle_update_goal(int(path.split("/")[3]), data)
            return self.send_error_json("Rota nao encontrada", HTTPStatus.NOT_FOUND)
        except ValueError as exc:
            return self.send_error_json(str(exc))
        except Exception as exc:
            return self.send_error_json(str(exc), HTTPStatus.INTERNAL_SERVER_ERROR)

    def do_DELETE(self):
        path, _ = parse_query(self.path)
        try:
            if self.require_user() is None:
                return
            if path.startswith("/api/users/"):
                return self.handle_delete_user(int(path.split("/")[3]))
            if path.startswith("/api/categories/"):
                return self.handle_delete_category(int(path.split("/")[3]))
            if path.startswith("/api/expenses/"):
                return self.handle_delete("expenses", int(path.split("/")[3]))
            if path.startswith("/api/incomes/"):
                return self.handle_delete("incomes", int(path.split("/")[3]))
            if path.startswith("/api/goals/"):
                return self.handle_delete("goals", int(path.split("/")[3]))
            return self.send_error_json("Rota nao encontrada", HTTPStatus.NOT_FOUND)
        except Exception as exc:
            return self.send_error_json(str(exc), HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_login(self, data):
        username = str(data.get("username", "")).strip()
        password = str(data.get("password", ""))
        with connect() as conn:
            user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
            if not user or not user["active"] or not verify_password(password, user["salt"], user["password_hash"]):
                return self.send_error_json("Usuario ou senha invalidos", HTTPStatus.UNAUTHORIZED)
            token = secrets.token_urlsafe(32)
            expires_at = (datetime.now() + timedelta(hours=SESSION_HOURS)).replace(microsecond=0).isoformat()
            conn.execute(
                "INSERT INTO sessions (token, user_id, expires_at) VALUES (?, ?, ?)",
                (token, user["id"], expires_at),
            )
            self.send_json({"token": token, "user": public_user(user)})

    def handle_register(self, data):
        payload = user_payload(data, creating=True, public=True)
        salt, password_hash = hash_password(payload["password"])
        try:
            with connect() as conn:
                cur = conn.execute(
                    """
                    INSERT INTO users (username, full_name, password_hash, salt, profile, active, created_at)
                    VALUES (?, ?, ?, ?, 'usuario', 1, ?)
                    """,
                    (payload["username"], payload["full_name"], password_hash, salt, now()),
                )
        except integrity_errors():
            return self.send_error_json("Usuario ja existe")
        self.send_json({"id": cur.lastrowid, "message": "Conta criada com sucesso. Faça login."}, HTTPStatus.CREATED)

    def handle_logout(self):
        token = self.headers.get("Authorization", "").replace("Bearer ", "").strip()
        with connect() as conn:
            conn.execute("DELETE FROM sessions WHERE token = ?", (token,))
        self.send_json({"ok": True})

    def handle_list_users(self):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            rows = conn.execute(
                "SELECT id, username, full_name, profile, active, created_at FROM users ORDER BY username"
            ).fetchall()
        self.send_json({"users": [public_user(row) for row in rows]})

    def handle_create_user(self, data):
        if self.require_superadmin() is None:
            return
        payload = user_payload(data, creating=True)
        salt, password_hash = hash_password(payload["password"])
        try:
            with connect() as conn:
                cur = conn.execute(
                    """
                    INSERT INTO users (username, full_name, password_hash, salt, profile, active, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        payload["username"],
                        payload["full_name"],
                        password_hash,
                        salt,
                        payload["profile"],
                        payload["active"],
                        now(),
                    ),
                )
        except integrity_errors():
            return self.send_error_json("Usuario ja existe")
        self.send_json({"id": cur.lastrowid}, HTTPStatus.CREATED)

    def handle_update_user(self, user_id, data):
        if self.require_superadmin() is None:
            return
        payload = user_payload(data)
        with connect() as conn:
            existing = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
            if not existing:
                return self.send_error_json("Usuario nao encontrado", HTTPStatus.NOT_FOUND)
            if existing["username"] == SUPERADMIN_USERNAME and payload["username"] != SUPERADMIN_USERNAME:
                return self.send_error_json("O SuperAdmin principal nao pode ser renomeado")
            profile = "superadmin" if existing["username"] == SUPERADMIN_USERNAME else payload["profile"]
            active = 1 if existing["username"] == SUPERADMIN_USERNAME else payload["active"]
            try:
                conn.execute(
                    "UPDATE users SET username = ?, full_name = ?, profile = ?, active = ? WHERE id = ?",
                    (payload["username"], payload["full_name"], profile, active, user_id),
                )
            except integrity_errors():
                return self.send_error_json("Usuario ja existe")
        self.send_json({"ok": True})

    def handle_admin_change_password(self, user_id, data):
        if self.require_superadmin() is None:
            return
        new_password = str(data.get("password", ""))
        if len(new_password) < 6:
            raise ValueError("A senha deve ter pelo menos 6 caracteres")
        salt, password_hash = hash_password(new_password)
        with connect() as conn:
            cur = conn.execute(
                "UPDATE users SET password_hash = ?, salt = ? WHERE id = ?",
                (password_hash, salt, user_id),
            )
        if cur.rowcount == 0:
            return self.send_error_json("Usuario nao encontrado", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True})

    def handle_toggle_user(self, user_id):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            user = conn.execute("SELECT username, active FROM users WHERE id = ?", (user_id,)).fetchone()
            if not user:
                return self.send_error_json("Usuario nao encontrado", HTTPStatus.NOT_FOUND)
            if user["username"] == SUPERADMIN_USERNAME:
                return self.send_error_json("O SuperAdmin principal nao pode ser desativado")
            new_active = 0 if user["active"] else 1
            conn.execute("UPDATE users SET active = ? WHERE id = ?", (new_active, user_id))
            if not new_active:
                conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
        self.send_json({"ok": True, "active": bool(new_active)})

    def handle_delete_user(self, user_id):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            user = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
            if not user:
                return self.send_error_json("Usuario nao encontrado", HTTPStatus.NOT_FOUND)
            if user["username"] == SUPERADMIN_USERNAME:
                return self.send_error_json("O SuperAdmin principal nao pode ser excluido")
            conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
            conn.execute("UPDATE users SET active = 0 WHERE id = ?", (user_id,))
        self.send_json({"success": True, "message": "Registro excluído com sucesso", "id": user_id, "deactivated": True})

    def handle_change_password(self, data):
        user = self.require_user()
        if not user:
            return
        old_password = str(data.get("old_password", ""))
        new_password = str(data.get("new_password", ""))
        if len(new_password) < 6:
            raise ValueError("A nova senha deve ter pelo menos 6 caracteres")
        with connect() as conn:
            row = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
            if not verify_password(old_password, row["salt"], row["password_hash"]):
                return self.send_error_json("Senha atual incorreta", HTTPStatus.UNAUTHORIZED)
            salt, password_hash = hash_password(new_password)
            conn.execute(
                "UPDATE users SET password_hash = ?, salt = ? WHERE id = ?",
                (password_hash, salt, user["id"]),
            )
        self.send_json({"ok": True})

    def handle_categories(self):
        if self.require_user() is None:
            return
        with connect() as conn:
            rows = conn.execute("SELECT name FROM categories WHERE active = 1 ORDER BY name").fetchall()
        self.send_json({"categories": [row["name"] for row in rows]})

    def handle_manage_categories(self):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            rows = conn.execute("SELECT id, name, active FROM categories ORDER BY name").fetchall()
        self.send_json({"categories": [row_to_dict(row) for row in rows]})

    def handle_create_category(self, data):
        if self.require_superadmin() is None:
            return
        name = str(data.get("name", "")).strip()
        if not name:
            raise ValueError("Nome da categoria obrigatorio")
        try:
            with connect() as conn:
                existing = conn.execute("SELECT id FROM categories WHERE name = ?", (name,)).fetchone()
                if existing:
                    conn.execute("UPDATE categories SET active = 1 WHERE id = ?", (existing["id"],))
                    return self.send_json({"id": existing["id"], "reactivated": True})
                cur = conn.execute("INSERT INTO categories (name, active) VALUES (?, 1)", (name,))
        except integrity_errors():
            return self.send_error_json("Categoria ja existe")
        self.send_json({"id": cur.lastrowid}, HTTPStatus.CREATED)

    def handle_update_category(self, category_id, data):
        if self.require_superadmin() is None:
            return
        name = str(data.get("name", "")).strip()
        active = 1 if str(data.get("active", "true")).lower() in ("1", "true", "sim", "on") else 0
        if not name:
            raise ValueError("Nome da categoria obrigatorio")
        try:
            with connect() as conn:
                cur = conn.execute(
                    "UPDATE categories SET name = ?, active = ? WHERE id = ?",
                    (name, active, category_id),
                )
        except integrity_errors():
            return self.send_error_json("Categoria ja existe")
        if cur.rowcount == 0:
            return self.send_error_json("Categoria nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True})

    def handle_delete_category(self, category_id):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            cur = conn.execute("UPDATE categories SET active = 0 WHERE id = ?", (category_id,))
        if cur.rowcount == 0:
            return self.send_error_json("Categoria nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"success": True, "message": "Registro excluído com sucesso", "id": category_id, "deactivated": True})

    def handle_maintenance_status(self):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            settings = maintenance_settings(conn)
            last = conn.execute("SELECT * FROM maintenance_cleanup_logs ORDER BY run_at DESC, id DESC LIMIT 1").fetchone()
            logs = conn.execute("SELECT * FROM maintenance_cleanup_logs ORDER BY run_at DESC, id DESC LIMIT 50").fetchall()
        self.send_json({
            "settings": settings,
            "last_cleanup": row_to_dict(last),
            "logs": [row_to_dict(row) for row in logs],
        })

    def handle_update_maintenance_settings(self, data):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            for key in DEFAULT_MAINTENANCE_SETTINGS:
                if key not in data:
                    continue
                try:
                    value = int(data.get(key))
                except (TypeError, ValueError):
                    raise ValueError("Dias de manutencao devem ser numeros inteiros")
                if value < 1 or value > 3650:
                    raise ValueError("Dias de manutencao devem ficar entre 1 e 3650")
                set_maintenance_setting(conn, key, value)
            settings = maintenance_settings(conn)
        self.send_json({"settings": settings})

    def handle_run_maintenance(self):
        if self.require_superadmin() is None:
            return
        with connect() as conn:
            result = run_cleanup(conn, "manual")
            settings = maintenance_settings(conn)
            logs = conn.execute("SELECT * FROM maintenance_cleanup_logs ORDER BY run_at DESC, id DESC LIMIT 50").fetchall()
        self.send_json({"result": result, "settings": settings, "logs": [row_to_dict(row) for row in logs]})

    def handle_dashboard(self, params):
        user = self.require_user()
        if user is None:
            return
        normalize_overdue_expenses()
        selected = dashboard_period(params)
        start, end = month_range(selected["year"], selected["month"])
        today = today_iso()
        scope_sql, scope_values = owned_update_suffix(user, params)
        with connect() as conn:
            expense_deleted_filter = not_deleted_filter(conn, "expenses")
            income_deleted_filter = not_deleted_filter(conn, "incomes")
            goal_deleted_filter = not_deleted_filter(conn, "goals")
            expenses = conn.execute(
                f"""
                SELECT * FROM expenses
                WHERE active = 1
                  AND cancelled = 0
                  AND status != 'Cancelada'
                  {expense_deleted_filter}
                  AND due_date >= ? AND due_date < ?{scope_sql}
                ORDER BY due_date
                """,
                (start, end, *scope_values),
            ).fetchall()
            incomes = conn.execute(
                f"""
                SELECT * FROM incomes
                WHERE active = 1
                  AND cancelled = 0
                  AND status != 'Cancelada'
                  {income_deleted_filter}
                  AND receipt_date >= ? AND receipt_date < ?{scope_sql}
                ORDER BY receipt_date
                """,
                (start, end, *scope_values),
            ).fetchall()
            overdue_rows = conn.execute(
                f"""
                SELECT * FROM expenses
                WHERE active = 1
                  AND status != 'Pago'
                  AND status != 'Cancelada'
                  AND cancelled = 0
                  {expense_deleted_filter}
                  AND due_date < ?{scope_sql}
                ORDER BY due_date
                """,
                (today, *scope_values),
            ).fetchall()
            upcoming_rows = conn.execute(
                f"""
                SELECT * FROM expenses
                WHERE active = 1
                  AND status != 'Pago'
                  AND status != 'Cancelada'
                  AND cancelled = 0
                  {expense_deleted_filter}
                  AND due_date >= ?{scope_sql}
                ORDER BY due_date
                """,
                (today, *scope_values),
            ).fetchall()
            goals = conn.execute(
                f"SELECT * FROM goals WHERE active = 1 AND status != 'Cancelada'{goal_deleted_filter}{scope_sql}",
                tuple(scope_values),
            ).fetchall()
            future_installments = conn.execute(
                f"""
                SELECT COALESCE(SUM(amount), 0) AS total
                FROM expenses
                WHERE active = 1
                  AND status != 'Pago'
                  AND status != 'Cancelada'
                  AND cancelled = 0
                  {expense_deleted_filter}
                  AND installment_group IS NOT NULL
                  AND installment_group != ''
                  AND installment_number IS NOT NULL
                  AND installment_total IS NOT NULL
                  AND due_date > ?{scope_sql}
                """,
                (today, *scope_values),
            ).fetchone()

        total_pending = sum(row["amount"] for row in expenses if row["status"] != "Pago")
        total_paid = sum(row["amount"] for row in expenses if row["status"] == "Pago")
        total_income = sum(row["amount"] for row in incomes)
        real_income = sum(row["amount"] for row in incomes if row["status"] == "Recebido")
        total_expenses = sum(row["amount"] for row in expenses)
        goals_total = sum(row["current_amount"] for row in goals)
        goals_target_total = sum(row["target_amount"] for row in goals)
        goals_progress = 0 if goals_target_total <= 0 else min(round((goals_total / goals_target_total) * 100, 2), 100)
        self.send_json(
            {
                "period": selected,
                "cards": {
                    "total_pending": round(total_pending, 2),
                    "total_paid": round(total_paid, 2),
                    "total_income": round(total_income, 2),
                    "total_expenses": round(total_expenses, 2),
                    "month_balance": round(total_income - total_expenses, 2),
                    "expected_balance": round(total_income - total_pending - total_paid, 2),
                    "real_balance": round(real_income - total_paid, 2),
                    "goals_total": round(goals_total, 2),
                    "goals_progress": goals_progress,
                    "future_installments": round(float(future_installments["total"] or 0), 2),
                    "overdue_count": len(overdue_rows),
                    "upcoming_count": len(upcoming_rows),
                },
                "overdue": [row_to_dict(row) for row in overdue_rows[:6]],
                "upcoming": [row_to_dict(row) for row in upcoming_rows[:6]],
            }
        )

    def handle_list_expenses(self, params):
        user = self.require_user()
        if user is None:
            return
        normalize_overdue_expenses()
        where, values = filters_to_sql(params, "expenses")
        if where:
            where += " AND active = 1 AND cancelled = 0 AND status != 'Cancelada'"
        else:
            where = " WHERE active = 1 AND cancelled = 0 AND status != 'Cancelada'"
        where, values = apply_user_scope(where, values, user, params)
        with connect() as conn:
            rows = conn.execute(f"SELECT * FROM expenses{where} ORDER BY due_date DESC, id DESC", values).fetchall()
        self.send_json({"expenses": [row_to_dict(row) for row in rows]})

    def handle_list_incomes(self, params):
        user = self.require_user()
        if user is None:
            return
        where, values = filters_to_sql(params, "incomes")
        where = (where + " AND active = 1 AND cancelled = 0 AND status != 'Cancelada'") if where else " WHERE active = 1 AND cancelled = 0 AND status != 'Cancelada'"
        where, values = apply_user_scope(where, values, user, params)
        with connect() as conn:
            rows = conn.execute(f"SELECT * FROM incomes{where} ORDER BY receipt_date DESC, id DESC", values).fetchall()
        self.send_json({"incomes": [row_to_dict(row) for row in rows]})

    def handle_list_goals(self, params):
        user = self.require_user()
        if user is None:
            return
        where = " WHERE active = 1 AND status != 'Cancelada'"
        values = []
        if user["profile"] == "superadmin" and params.get("user_id"):
            where += " AND user_id = ?"
            values.append(params.get("user_id"))
        elif user["profile"] != "superadmin":
            where += " AND user_id = ?"
            values.append(user["id"])
        with connect() as conn:
            rows = conn.execute(f"SELECT * FROM goals{where} ORDER BY target_date, id", values).fetchall()
        self.send_json({"goals": [row_to_dict(row) for row in rows]})

    def handle_create_expense(self, data):
        user = self.require_user()
        if user is None:
            return
        payload = expense_payload(data)
        stamp = now()
        is_installment = is_yes(data.get("is_installment", "Nao"))
        installment_total = clean_int(data.get("installment_total"), 1) if is_installment else 1
        first_due_date = str(data.get("first_due_date") or payload["due_date"]).strip()
        if installment_total < 1:
            raise ValueError("Quantidade de parcelas invalida")
        installment_group = secrets.token_hex(8) if is_installment else None
        created_ids = []
        with connect() as conn:
            for index in range(installment_total):
                installment_number = index + 1 if is_installment else None
                description = payload["description"]
                due_date = payload["due_date"]
                if is_installment:
                    description = f"{payload['description']} ({index + 1}/{installment_total})"
                    due_date = add_months(first_due_date, index)
                amount = payload["amount"]
                cur = conn.execute(
                    """
                    INSERT INTO expenses
                    (description, category, amount, due_date, status, payment_method, notes, payment_date, user_id,
                     installment_group, installment_number, installment_total, cancelled, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                    """,
                    (
                        description,
                        payload["category"],
                        amount,
                        due_date,
                        payload["status"],
                        payload["payment_method"],
                        payload["notes"],
                        payload["payment_date"],
                        user["id"],
                        installment_group,
                        installment_number,
                        installment_total if is_installment else None,
                        stamp,
                        stamp,
                    ),
                )
                created_ids.append(cur.lastrowid)
        self.send_json({"id": created_ids[0], "ids": created_ids, "installments": len(created_ids)}, HTTPStatus.CREATED)

    def handle_update_expense(self, expense_id, data):
        user = self.require_user()
        if user is None:
            return
        payload = expense_payload(data)
        update_scope = str(data.get("update_scope", "single")).strip()
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            affected = update_expense_records(conn, expense_id, payload, update_scope, scope_sql, scope_values)
        if affected == 0:
            return self.send_error_json("Conta nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True, "updated": affected})

    def handle_pay_expense(self, expense_id, data):
        user = self.require_user()
        if user is None:
            return
        payment_date = str(data.get("payment_date") or today_iso()).strip()
        payment_method = str(data.get("payment_method", "")).strip()
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            cur = conn.execute(
                f"""
                UPDATE expenses
                SET status = 'Pago', payment_date = ?, payment_method = COALESCE(NULLIF(?, ''), payment_method), updated_at = ?
                WHERE id = ?{scope_sql}
                """,
                (payment_date, payment_method, now(), expense_id, *scope_values),
            )
        if cur.rowcount == 0:
            return self.send_error_json("Conta nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True})

    def handle_cancel_future_installments(self, expense_id):
        user = self.require_user()
        if user is None:
            return
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            row = conn.execute(
                f"SELECT installment_group, installment_number FROM expenses WHERE id = ?{scope_sql}",
                (expense_id, *scope_values),
            ).fetchone()
            if not row or not row["installment_group"]:
                return self.send_error_json("Despesa parcelada nao encontrada", HTTPStatus.NOT_FOUND)
            cur = conn.execute(
                f"""
                UPDATE expenses
                SET cancelled = 1, updated_at = ?
                WHERE installment_group = ? AND status != 'Pago' AND installment_number > ?{scope_sql}
                """,
                (now(), row["installment_group"], row["installment_number"], *scope_values),
            )
        self.send_json({"ok": True, "cancelled": cur.rowcount})

    def handle_create_income(self, data):
        user = self.require_user()
        if user is None:
            return
        payload = income_payload(data)
        stamp = now()
        is_recurring = is_yes(data.get("is_recurring") or data.get("recorrente"))
        recurring_total = clean_int(data.get("recurring_total") or data.get("quantidade_parcelas"), 1) if is_recurring else 1
        first_receipt_date = str(data.get("first_receipt_date") or data.get("data_primeiro_recebimento") or payload["receipt_date"]).strip()
        amount_per_item = clean_money(data.get("recurring_amount") or data.get("valor_parcela") or payload["amount"]) if is_recurring else payload["amount"]
        if recurring_total < 1:
            raise ValueError("Quantidade de meses/parcelas invalida")
        recurring_group = secrets.token_hex(8) if is_recurring else None
        created_ids = []
        with connect() as conn:
            for index in range(recurring_total):
                recurring_number = index + 1 if is_recurring else None
                description = payload["description"]
                receipt_date = payload["receipt_date"]
                if is_recurring:
                    description = f"{payload['description']} ({index + 1}/{recurring_total})"
                    receipt_date = add_months(first_receipt_date, index)
                cur = conn.execute(
                    """
                    INSERT INTO incomes
                    (description, category, amount, receipt_date, status, notes, user_id,
                     recurring_group, recurring_number, recurring_total, cancelled, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                    """,
                    (
                        description,
                        payload["category"],
                        amount_per_item,
                        receipt_date,
                        payload["status"],
                        payload["notes"],
                        user["id"],
                        recurring_group,
                        recurring_number,
                        recurring_total if is_recurring else None,
                        stamp,
                        stamp,
                    ),
                )
                created_ids.append(cur.lastrowid)
        self.send_json({"id": created_ids[0], "ids": created_ids, "recurrences": len(created_ids)}, HTTPStatus.CREATED)

    def handle_update_income(self, income_id, data):
        user = self.require_user()
        if user is None:
            return
        payload = income_payload(data)
        update_scope = str(data.get("update_scope", "single")).strip()
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            affected = update_income_records(conn, income_id, payload, update_scope, scope_sql, scope_values)
        if affected == 0:
            return self.send_error_json("Receita nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True, "updated": affected})

    def handle_receive_income(self, income_id):
        user = self.require_user()
        if user is None:
            return
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            cur = conn.execute(
                f"UPDATE incomes SET status = 'Recebido', updated_at = ? WHERE id = ?{scope_sql}",
                (now(), income_id, *scope_values),
            )
        if cur.rowcount == 0:
            return self.send_error_json("Receita nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True})

    def handle_cancel_future_incomes(self, income_id):
        user = self.require_user()
        if user is None:
            return
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            row = conn.execute(
                f"SELECT recurring_group, recurring_number FROM incomes WHERE id = ?{scope_sql}",
                (income_id, *scope_values),
            ).fetchone()
            if not row or not row["recurring_group"]:
                return self.send_error_json("Receita recorrente nao encontrada", HTTPStatus.NOT_FOUND)
            cur = conn.execute(
                f"""
                UPDATE incomes
                SET cancelled = 1, updated_at = ?
                WHERE recurring_group = ? AND status != 'Recebido' AND recurring_number > ?{scope_sql}
                """,
                (now(), row["recurring_group"], row["recurring_number"], *scope_values),
            )
        self.send_json({"ok": True, "cancelled": cur.rowcount})

    def handle_list_goals(self, params):
        user = self.require_user()
        if user is None:
            return
        where, values = apply_user_scope(" WHERE active = 1 AND status != 'Cancelada'", [], user, params)
        with connect() as conn:
            rows = conn.execute(f"SELECT * FROM goals{where} ORDER BY target_date, id DESC", values).fetchall()
        self.send_json({"goals": [enrich_goal(row) for row in rows]})

    def handle_create_goal(self, data):
        user = self.require_user()
        if user is None:
            return
        payload = goal_payload(data)
        stamp = now()
        with connect() as conn:
            cur = conn.execute(
                """
                INSERT INTO goals
                (user_id, name, description, target_amount, current_amount, target_date, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    user["id"],
                    payload["name"],
                    payload["description"],
                    payload["target_amount"],
                    payload["current_amount"],
                    payload["target_date"],
                    payload["status"],
                    stamp,
                    stamp,
                ),
            )
        self.send_json({"id": cur.lastrowid}, HTTPStatus.CREATED)

    def handle_update_goal(self, goal_id, data):
        user = self.require_user()
        if user is None:
            return
        payload = goal_payload(data)
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            cur = conn.execute(
                f"""
                UPDATE goals
                SET name = ?, description = ?, target_amount = ?, current_amount = ?,
                    target_date = ?, status = ?, updated_at = ?
                WHERE id = ?{scope_sql}
                """,
                (
                    payload["name"],
                    payload["description"],
                    payload["target_amount"],
                    payload["current_amount"],
                    payload["target_date"],
                    payload["status"],
                    now(),
                    goal_id,
                    *scope_values,
                ),
            )
        if cur.rowcount == 0:
            return self.send_error_json("Meta nao encontrada", HTTPStatus.NOT_FOUND)
        self.send_json({"ok": True})

    def handle_goal_amount(self, goal_id, data, operation):
        user = self.require_user()
        if user is None:
            return
        amount = clean_money(data.get("amount", 0))
        scope_sql, scope_values = owned_update_suffix(user)
        sign = 1 if operation == "add" else -1
        with connect() as conn:
            row = conn.execute(f"SELECT current_amount FROM goals WHERE active = 1 AND id = ?{scope_sql}", (goal_id, *scope_values)).fetchone()
            if not row:
                return self.send_error_json("Meta nao encontrada", HTTPStatus.NOT_FOUND)
            new_amount = max(float(row["current_amount"]) + (amount * sign), 0)
            conn.execute("UPDATE goals SET current_amount = ?, updated_at = ? WHERE id = ?", (round(new_amount, 2), now(), goal_id))
        self.send_json({"ok": True, "current_amount": round(new_amount, 2)})

    def handle_delete(self, table, row_id):
        user = self.require_user()
        if user is None:
            return
        scope_sql, scope_values = owned_update_suffix(user)
        with connect() as conn:
            cur = conn.execute(f"UPDATE {table} SET active = 0, updated_at = ? WHERE id = ?{scope_sql}", (now(), row_id, *scope_values))
        if cur.rowcount == 0:
            return self.send_error_json("Registro nao encontrado", HTTPStatus.NOT_FOUND)
        self.send_json({"success": True, "message": "Registro excluído com sucesso", "id": row_id, "table": table})

    def handle_report(self, params):
        user = self.require_user()
        if user is None:
            return
        data = build_report(params, user)
        self.send_json(data)

    def handle_report_export(self, params):
        user = self.require_user()
        if user is None:
            return
        report = build_report(params, user)
        rows = [
            ["Resumo", "Valor"],
            ["Total de receitas", money_value(report["summary"]["total_income"])],
            ["Total de despesas", money_value(report["summary"]["total_expense"])],
            ["Saldo final", money_value(report["summary"]["final_balance"])],
            [],
            ["Tipo", "Descricao", "Categoria", "Valor", "Data", "Status", "Forma de pagamento", "Observacao"],
        ]
        for item in report["paid_expenses"]:
            rows.append(["Despesa paga", item["description"], item["category"], item["amount"], item["due_date"], item["status"], item.get("payment_method", ""), item.get("notes", "")])
        for item in report["pending_expenses"]:
            rows.append(["Despesa pendente", item["description"], item["category"], item["amount"], item["due_date"], item["status"], item.get("payment_method", ""), item.get("notes", "")])
        for item in report["incomes"]:
            rows.append(["Receita", item["description"], item["category"], item["amount"], item["receipt_date"], item["status"], "", item.get("notes", "")])

        text = csv_string(rows)
        payload = text.encode("utf-8-sig")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", "attachment; filename=relatorio-financeiro.csv")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def handle_charts(self, params):
        user = self.require_user()
        if user is None:
            return
        selected_year = int(params.get("year") or date.today().year)
        months = [f"{selected_year}-{month:02d}" for month in range(1, 13)]
        scope_sql, scope_values = owned_update_suffix(user, params)
        with connect() as conn:
            expense_deleted_filter = not_deleted_filter(conn, "expenses")
            income_deleted_filter = not_deleted_filter(conn, "incomes")
            goal_deleted_filter = not_deleted_filter(conn, "goals")
            expenses = conn.execute(
                f"""
                SELECT substr(due_date, 1, 7) AS month, COALESCE(SUM(amount), 0) AS total
                FROM expenses
                WHERE active = 1
                  AND cancelled = 0
                  AND status != 'Cancelada'
                  {expense_deleted_filter}
                  AND due_date >= ? AND due_date < ?{scope_sql}
                GROUP BY substr(due_date, 1, 7)
                """,
                (months[0] + "-01", f"{selected_year + 1}-01-01", *scope_values),
            ).fetchall()
            incomes = conn.execute(
                f"""
                SELECT substr(receipt_date, 1, 7) AS month, COALESCE(SUM(amount), 0) AS total
                FROM incomes
                WHERE active = 1
                  AND cancelled = 0
                  AND status != 'Cancelada'
                  {income_deleted_filter}
                  AND receipt_date >= ? AND receipt_date < ?{scope_sql}
                GROUP BY substr(receipt_date, 1, 7)
                """,
                (months[0] + "-01", f"{selected_year + 1}-01-01", *scope_values),
            ).fetchall()
            goals = conn.execute(
                f"SELECT COALESCE(SUM(current_amount), 0) AS total FROM goals WHERE active = 1 AND status != 'Cancelada'{goal_deleted_filter}{scope_sql}",
                tuple(scope_values),
            ).fetchone()
        expense_map = {row["month"]: row["total"] for row in expenses}
        income_map = {row["month"]: row["total"] for row in incomes}
        self.send_json(
            {
                "months": months,
                "incomes": [round(float(income_map.get(month, 0)), 2) for month in months],
                "expenses": [round(float(expense_map.get(month, 0)), 2) for month in months],
                "balances": [round(float(income_map.get(month, 0)) - float(expense_map.get(month, 0)), 2) for month in months],
                "goals": [round(float(goals["total"] or 0), 2) for _ in months],
            }
        )

    def handle_cashflow(self, params):
        user = self.require_user()
        if user is None:
            return
        months = list(iter_months(date.today().replace(day=1), 12))
        scope_sql, scope_values = owned_update_suffix(user, params)
        rows = []
        with connect() as conn:
            expense_deleted_filter = not_deleted_filter(conn, "expenses")
            income_deleted_filter = not_deleted_filter(conn, "incomes")
            goal_deleted_filter = not_deleted_filter(conn, "goals")
            for month in months:
                year, month_num = map(int, month.split("-"))
                start, end = month_range(year, month_num)
                income = conn.execute(
                    f"SELECT COALESCE(SUM(amount), 0) AS total FROM incomes WHERE active = 1 AND cancelled = 0 AND status != 'Cancelada'{income_deleted_filter} AND receipt_date >= ? AND receipt_date < ?{scope_sql}",
                    (start, end, *scope_values),
                ).fetchone()["total"]
                expense = conn.execute(
                    f"SELECT COALESCE(SUM(amount), 0) AS total FROM expenses WHERE active = 1 AND cancelled = 0 AND status != 'Cancelada'{expense_deleted_filter} AND due_date >= ? AND due_date < ?{scope_sql}",
                    (start, end, *scope_values),
                ).fetchone()["total"]
                goals = conn.execute(
                    f"SELECT COALESCE(SUM(current_amount), 0) AS total FROM goals WHERE active = 1 AND status != 'Cancelada'{goal_deleted_filter} AND substr(updated_at, 1, 7) = ?{scope_sql}",
                    (month, *scope_values),
                ).fetchone()["total"]
                rows.append(
                    {
                        "month": month,
                        "incomes": round(float(income or 0), 2),
                        "expenses": round(float(expense or 0), 2),
                        "goals": round(float(goals or 0), 2),
                        "balance": round(float(income or 0) - float(expense or 0), 2),
                    }
                )
        self.send_json({"cashflow": rows})

    def handle_import_template(self):
        rows = [
            ["tipo", "descricao_ou_nome", "categoria", "valor", "data", "status", "observacao"],
            ["despesa", "Consulta", "Saúde", "200.00", "2026-06-15", "Pendente", "Especialidade"],
            ["receita", "Salario", "Salario", "3500.00", "2026-06-05", "Recebido", ""],
            ["meta", "Reserva de Emergencia", "", "10000.00", "2026-12-31", "Ativa", "Meta principal"],
        ]
        payload = csv_string(rows).encode("utf-8-sig")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", "attachment; filename=modelo-importacao-financeira.csv")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def validate_import_rows_legacy_csv(self, rows):
        valid = []
        errors = []
        if not isinstance(rows, list):
            raise ValueError("Linhas invalidas")
        for index, row in enumerate(rows, start=1):
            if not isinstance(row, dict) or not any(str(value).strip() for value in row.values()):
                continue
            item_type = str(row.get("tipo", "")).strip().lower()
            description = str(row.get("descricao_ou_nome", "")).strip()
            category = str(row.get("categoria", "")).strip()
            date_value = str(row.get("data", "")).strip()
            status = str(row.get("status", "")).strip()
            try:
                amount = clean_money(row.get("valor", 0))
            except ValueError as exc:
                errors.append({"line": index, "error": str(exc)})
                continue
            if item_type not in ("despesa", "receita"):
                errors.append({"line": index, "error": "Tipo invalido"})
                continue
            if not description or not category or not date_value:
                errors.append({"line": index, "error": "Descricao, categoria e data sao obrigatorias"})
                continue
            if item_type == "despesa" and status not in ("Pendente", "Pago", "Vencido"):
                errors.append({"line": index, "error": "Status de despesa invalido"})
                continue
            if item_type == "receita" and status not in ("Recebido", "Pendente"):
                errors.append({"line": index, "error": "Status de receita invalido"})
                continue
            valid.append({**row, "tipo": item_type, "descricao_ou_nome": description, "categoria": category, "valor": amount, "data": date_value, "status": status})
        return valid, errors

    def handle_import_preview_legacy_csv(self, data):
        if self.require_user() is None:
            return
        valid, errors = self.validate_import_rows_legacy_csv(data.get("rows", []))
        self.send_json({"valid": valid, "errors": errors})

    def handle_import_commit_legacy_csv(self, data):
        user = self.require_user()
        if user is None:
            return
        target_user_id = user["id"]
        if user["profile"] == "superadmin" and data.get("user_id"):
            target_user_id = int(data.get("user_id"))
        valid, errors = self.validate_import_rows_legacy_csv(data.get("rows", []))
        if errors:
            return self.send_json({"imported": 0, "errors": errors}, HTTPStatus.BAD_REQUEST)
        stamp = now()
        imported = 0
        with connect() as conn:
            for row in valid:
                if row["tipo"] == "despesa":
                    conn.execute(
                        """
                        INSERT INTO expenses
                        (description, category, amount, due_date, status, notes, user_id, active, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                        """,
                        (row["descricao_ou_nome"], row["categoria"], row["valor"], row["data"], row["status"], row.get("observacao", ""), target_user_id, stamp, stamp),
                    )
                else:
                    conn.execute(
                        """
                        INSERT INTO incomes
                        (description, category, amount, receipt_date, status, notes, user_id, active, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                        """,
                        (row["descricao_ou_nome"], row["categoria"], row["valor"], row["data"], row["status"], row.get("observacao", ""), target_user_id, stamp, stamp),
                    )
                imported += 1
        self.send_json({"imported": imported, "errors": []})

    def handle_import_template(self):
        payload = build_import_template()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", XLSX_CONTENT_TYPE)
        self.send_header("Content-Disposition", "attachment; filename=modelo-importacao-financeira.xlsx")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def decode_import_workbook(self, data):
        filename = str(data.get("filename", "")).strip()
        content = str(data.get("content_base64", ""))
        print(f"[importacao] decode XLSX solicitado. arquivo={filename or '(sem nome)'} base64_chars={len(content)}")
        if not content:
            raise ValueError("Arquivo XLSX obrigatorio")
        if "," in content:
            content = content.split(",", 1)[1]
        try:
            payload = base64.b64decode(content, validate=True)
        except Exception as exc:
            print(f"[importacao] erro ao decodificar base64: {type(exc).__name__}: {exc}")
            raise ValueError("Arquivo XLSX invalido. O conteudo recebido nao esta em base64 valido.") from exc
        print(f"[importacao] XLSX decodificado. bytes={len(payload)} abas_esperadas={list(IMPORT_SHEETS.keys())}")
        return read_xlsx_workbook(payload)

    def validate_import_workbook(self, workbook):
        valid = []
        errors = []
        if not any(name in workbook for name in IMPORT_SHEETS):
            errors.append({"sheet": "ARQUIVO", "line": 0, "error": "Nenhuma aba oficial encontrada"})
            return valid, errors
        for sheet_name, config in IMPORT_SHEETS.items():
            rows = workbook.get(sheet_name)
            if not rows:
                continue
            header_index = next((idx for idx, row in enumerate(rows) if any(str(cell).strip() for cell in row)), None)
            if header_index is None:
                continue
            headers = [canonical_import_header(config, cell) for cell in rows[header_index]]
            positions = {}
            for index, header in enumerate(headers):
                if header and header not in positions:
                    positions[header] = index
            print(f"[importacao] aba {sheet_name}: cabecalhos normalizados={headers}")
            missing = [header for header in config.get("required_headers", config["headers"]) if header not in positions]
            if missing:
                errors.append({"sheet": sheet_name, "line": header_index + 1, "error": "Colunas obrigatorias ausentes: " + ", ".join(missing)})
                continue
            for row_number, values in enumerate(rows[header_index + 1:], start=header_index + 2):
                mapped = {
                    header: values[positions[header]] if header in positions and positions[header] < len(values) else ""
                    for header in config["headers"]
                }
                if not any(str(value).strip() for value in mapped.values()):
                    continue
                try:
                    valid.append(self.normalize_import_row(sheet_name, row_number, mapped))
                except ValueError as exc:
                    errors.append({"sheet": sheet_name, "line": row_number, "error": str(exc)})
        return valid, errors

    def normalize_import_row(self, sheet_name, row_number, row):
        if sheet_name == "DESPESAS":
            description = str(row.get("descricao", "")).strip()
            category = str(row.get("categoria", "")).strip()
            status = str(row.get("status", "") or "Pendente").strip()
            if not description or not category:
                raise ValueError("Descricao e categoria sao obrigatorias")
            if status not in ("Pendente", "Pago", "Vencido"):
                raise ValueError("Status de despesa invalido")
            return {"sheet": sheet_name, "line": row_number, "tipo": "despesa", "description": description, "category": category, "amount": clean_money(row.get("valor")), "due_date": clean_date(row.get("data_vencimento")), "status": status, "payment_method": str(row.get("forma_pagamento", "")).strip(), "notes": str(row.get("observacao", "")).strip()}
        if sheet_name == "RECEITAS":
            description = str(row.get("descricao", "")).strip()
            category = str(row.get("categoria", "")).strip()
            status = str(row.get("status", "") or "Pendente").strip()
            if not description or not category:
                raise ValueError("Descricao e categoria sao obrigatorias")
            if status not in ("Recebido", "Pendente"):
                raise ValueError("Status de receita invalido")
            recurring = is_yes(row.get("recorrente"))
            total = clean_int(row.get("quantidade_parcelas"), 1) if recurring else 1
            if recurring and total <= 0:
                raise ValueError("Quantidade de parcelas deve ser maior que zero")
            return {
                "sheet": sheet_name,
                "line": row_number,
                "tipo": "receita",
                "description": description,
                "category": category,
                "amount": clean_money(row.get("valor")),
                "receipt_date": clean_date(row.get("data_recebimento")),
                "status": status,
                "notes": str(row.get("observacao", "")).strip(),
                "recurring": recurring,
                "recurring_total": total,
                "first_receipt_date": clean_date(row.get("data_primeiro_recebimento") or row.get("data_recebimento")) if recurring else "",
            }
        if sheet_name == "METAS":
            name = str(row.get("nome", "")).strip()
            if not name:
                raise ValueError("Nome da meta obrigatorio")
            return {"sheet": sheet_name, "line": row_number, "tipo": "meta", "name": name, "description": str(row.get("descricao", "")).strip(), "target_amount": clean_money(row.get("valor_objetivo")), "current_amount": clean_money(row.get("valor_atual") or 0), "target_date": clean_date(row.get("data_prevista")), "status": str(row.get("status", "") or "Em andamento").strip()}
        if sheet_name == "PARCELADAS":
            description = str(row.get("descricao", "")).strip()
            category = str(row.get("categoria", "")).strip()
            status = str(row.get("status", "") or "Pendente").strip()
            if not description or not category:
                raise ValueError("Descricao e categoria sao obrigatorias")
            if status not in ("Pendente", "Pago", "Vencido"):
                raise ValueError("Status de parcela invalido")
            total = int(float(str(row.get("quantidade_parcelas", "")).replace(",", ".") or 0))
            if total <= 0:
                raise ValueError("Quantidade de parcelas deve ser maior que zero")
            return {
                "sheet": sheet_name,
                "line": row_number,
                "tipo": "parcelada",
                "description": description,
                "category": category,
                "installment_amount": clean_money(row.get("valor_parcela")),
                "installment_total": total,
                "first_due_date": clean_date(row.get("data_primeira_parcela")),
                "status": status,
                "payment_method": str(row.get("forma_pagamento", "")).strip(),
                "notes": str(row.get("observacao", "")).strip(),
            }
        raise ValueError("Aba nao suportada")

    def handle_import_preview(self, data):
        if self.require_user() is None:
            return
        workbook = self.decode_import_workbook(data)
        valid, errors = self.validate_import_workbook(workbook)
        self.send_json({"valid": valid, "errors": errors})

    def handle_import_commit(self, data):
        user = self.require_user()
        if user is None:
            return
        target_user_id = user["id"]
        if user["profile"] == "superadmin" and data.get("user_id"):
            target_user_id = int(data.get("user_id"))
        if data.get("rows"):
            valid, errors = data.get("rows", []), []
        else:
            workbook = self.decode_import_workbook(data)
            valid, errors = self.validate_import_workbook(workbook)
        if errors:
            return self.send_json({"imported": 0, "skipped": 0, "errors": errors}, HTTPStatus.BAD_REQUEST)
        stamp = now()
        imported = 0
        skipped = 0
        with connect() as conn:
            for row in valid:
                created = insert_import_row(conn, row, target_user_id, stamp)
                imported += created
                skipped += 0 if created else 1
            conn.execute(
                """
                INSERT INTO import_history
                (user_id, filename, imported_count, skipped_count, errors_count, is_test, created_at)
                VALUES (?, ?, ?, ?, 0, ?, ?)
                """,
                (
                    target_user_id,
                    str(data.get("filename", "")).strip(),
                    imported,
                    skipped,
                    1 if data.get("is_test") else 0,
                    stamp,
                ),
            )
        self.send_json({"imported": imported, "skipped": skipped, "errors": []})


def group_update_rows(conn, table, id_value, group_field, number_field, scope_sql, scope_values, update_scope):
    current = conn.execute(
        f"SELECT * FROM {table} WHERE id = ?{scope_sql}",
        (id_value, *scope_values),
    ).fetchone()
    if not current:
        return current, []
    group = current[group_field]
    if update_scope == "single" or not group:
        return current, [current]
    where = f"{group_field} = ?"
    values = [group]
    if update_scope == "future":
        where += f" AND {number_field} >= ?"
        values.append(current[number_field])
    elif update_scope != "all":
        return current, [current]
    if scope_sql:
        where += scope_sql.replace(" AND ", " AND ", 1)
        values.extend(scope_values)
    rows = conn.execute(f"SELECT * FROM {table} WHERE {where} ORDER BY {number_field}, id", values).fetchall()
    return current, rows


def update_expense_records(conn, expense_id, payload, update_scope, scope_sql, scope_values):
    current, rows = group_update_rows(conn, "expenses", expense_id, "installment_group", "installment_number", scope_sql, scope_values, update_scope)
    if not rows:
        return 0
    stamp = now()
    base_description = base_installment_description(payload["description"])
    for row in rows:
        description = payload["description"]
        due_date = payload["due_date"]
        if update_scope != "single" and current["installment_group"]:
            description = f"{base_description} ({row['installment_number']}/{row['installment_total']})"
            due_date = add_months(payload["due_date"], int(row["installment_number"]) - int(current["installment_number"]))
        conn.execute(
            f"""
            UPDATE expenses
            SET description = ?, category = ?, amount = ?, due_date = ?, status = ?,
                payment_method = ?, notes = ?, payment_date = ?, updated_at = ?
            WHERE id = ?{scope_sql}
            """,
            (
                description,
                payload["category"],
                payload["amount"],
                due_date,
                payload["status"],
                payload["payment_method"],
                payload["notes"],
                payload["payment_date"],
                stamp,
                row["id"],
                *scope_values,
            ),
        )
    return len(rows)


def update_income_records(conn, income_id, payload, update_scope, scope_sql, scope_values):
    current, rows = group_update_rows(conn, "incomes", income_id, "recurring_group", "recurring_number", scope_sql, scope_values, update_scope)
    if not rows:
        return 0
    stamp = now()
    base_description = base_installment_description(payload["description"])
    for row in rows:
        description = payload["description"]
        receipt_date = payload["receipt_date"]
        if update_scope != "single" and current["recurring_group"]:
            description = f"{base_description} ({row['recurring_number']}/{row['recurring_total']})"
            receipt_date = add_months(payload["receipt_date"], int(row["recurring_number"]) - int(current["recurring_number"]))
        conn.execute(
            f"""
            UPDATE incomes
            SET description = ?, category = ?, amount = ?, receipt_date = ?, status = ?,
                notes = ?, updated_at = ?
            WHERE id = ?{scope_sql}
            """,
            (
                description,
                payload["category"],
                payload["amount"],
                receipt_date,
                payload["status"],
                payload["notes"],
                stamp,
                row["id"],
                *scope_values,
            ),
        )
    return len(rows)


def duplicate_exists(conn, table, where, values):
    row = conn.execute(f"SELECT id FROM {table} WHERE active = 1 AND {where} LIMIT 1", values).fetchone()
    return row is not None


def insert_import_row(conn, row, user_id, stamp):
    kind = row.get("tipo")
    if kind == "despesa":
        if duplicate_exists(
            conn,
            "expenses",
            "user_id = ? AND description = ? AND category = ? AND amount = ? AND due_date = ?",
            (user_id, row["description"], row["category"], row["amount"], row["due_date"]),
        ):
            return 0
        conn.execute(
            """
            INSERT INTO expenses
            (description, category, amount, due_date, status, payment_method, notes, user_id, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (row["description"], row["category"], row["amount"], row["due_date"], row["status"], row.get("payment_method", ""), row.get("notes", ""), user_id, stamp, stamp),
        )
        return 1
    if kind == "receita":
        if row.get("recurring"):
            total = int(row.get("recurring_total") or 1)
            group = hashlib.sha256(f'{user_id}|{row["description"]}|{row["category"]}|{row["amount"]}|{row["first_receipt_date"]}|{total}'.encode("utf-8")).hexdigest()[:16]
            created = 0
            for number in range(1, total + 1):
                receipt_date = add_months(row["first_receipt_date"], number - 1)
                description = f'{row["description"]} ({number}/{total})'
                if duplicate_exists(
                    conn,
                    "incomes",
                    "user_id = ? AND description = ? AND category = ? AND amount = ? AND receipt_date = ?",
                    (user_id, description, row["category"], row["amount"], receipt_date),
                ):
                    continue
                conn.execute(
                    """
                    INSERT INTO incomes
                    (description, category, amount, receipt_date, status, notes, user_id, active,
                     recurring_group, recurring_number, recurring_total, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)
                    """,
                    (description, row["category"], row["amount"], receipt_date, row["status"], row.get("notes", ""), user_id, group, number, total, stamp, stamp),
                )
                created += 1
            return created
        if duplicate_exists(
            conn,
            "incomes",
            "user_id = ? AND description = ? AND category = ? AND amount = ? AND receipt_date = ?",
            (user_id, row["description"], row["category"], row["amount"], row["receipt_date"]),
        ):
            return 0
        conn.execute(
            """
            INSERT INTO incomes
            (description, category, amount, receipt_date, status, notes, user_id, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (row["description"], row["category"], row["amount"], row["receipt_date"], row["status"], row.get("notes", ""), user_id, stamp, stamp),
        )
        return 1
    if kind == "meta":
        if duplicate_exists(
            conn,
            "goals",
            "user_id = ? AND name = ? AND target_amount = ? AND target_date = ?",
            (user_id, row["name"], row["target_amount"], row["target_date"]),
        ):
            return 0
        conn.execute(
            """
            INSERT INTO goals
            (user_id, name, description, target_amount, current_amount, target_date, status, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (user_id, row["name"], row.get("description", ""), row["target_amount"], row["current_amount"], row["target_date"], row["status"], stamp, stamp),
        )
        return 1
    if kind == "parcelada":
        total = int(row["installment_total"])
        amount = round(row["installment_amount"], 2)
        group = hashlib.sha256(f'{user_id}|{row["description"]}|{row["category"]}|{amount}|{row["first_due_date"]}|{total}|{stamp}'.encode("utf-8")).hexdigest()[:16]
        created = 0
        for number in range(1, total + 1):
            due_date = add_months(row["first_due_date"], number - 1)
            description = f'{row["description"]} ({number}/{total})'
            if duplicate_exists(
                conn,
                "expenses",
                "user_id = ? AND description = ? AND category = ? AND amount = ? AND due_date = ?",
                (user_id, description, row["category"], amount, due_date),
            ):
                continue
            conn.execute(
                """
                INSERT INTO expenses
                (description, category, amount, due_date, status, payment_method, notes, user_id, active, installment_group, installment_number, installment_total, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)
                """,
                (description, row["category"], amount, due_date, row["status"], row.get("payment_method", ""), row.get("notes", ""), user_id, group, number, total, stamp, stamp),
            )
            created += 1
        return created
    return 0


def dashboard_period(params):
    today = date.today()
    return {
        "month": int(params.get("month") or today.month),
        "year": int(params.get("year") or today.year),
    }


def build_report(params, user):
    normalize_overdue_expenses()
    selected = dashboard_period(params)
    start, end = month_range(selected["year"], selected["month"])
    movement_type = params.get("type", "")
    scope_sql, scope_values = owned_update_suffix(user, params)
    with connect() as conn:
        expense_deleted_filter = not_deleted_filter(conn, "expenses")
        income_deleted_filter = not_deleted_filter(conn, "incomes")
        expenses = [] if movement_type == "Receita" else [row_to_dict(row) for row in conn.execute(
            f"SELECT * FROM expenses WHERE active = 1 AND cancelled = 0 AND status != 'Cancelada'{expense_deleted_filter} AND due_date >= ? AND due_date < ?{scope_sql} ORDER BY due_date, id",
            (start, end, *scope_values),
        ).fetchall()]
        incomes = [] if movement_type == "Despesa" else [row_to_dict(row) for row in conn.execute(
            f"SELECT * FROM incomes WHERE active = 1 AND cancelled = 0 AND status != 'Cancelada'{income_deleted_filter} AND receipt_date >= ? AND receipt_date < ?{scope_sql} ORDER BY receipt_date, id",
            (start, end, *scope_values),
        ).fetchall()]

    total_income = sum(item["amount"] for item in incomes)
    total_expense = sum(item["amount"] for item in expenses)
    paid_expenses = [item for item in expenses if item["status"] == "Pago"]
    pending_expenses = [item for item in expenses if item["status"] != "Pago"]
    return {
        "period": selected,
        "summary": {
            "total_income": round(total_income, 2),
            "total_expense": round(total_expense, 2),
            "final_balance": round(total_income - total_expense, 2),
        },
        "incomes": incomes,
        "paid_expenses": paid_expenses,
        "pending_expenses": pending_expenses,
    }


def money_value(value):
    return f"{float(value):.2f}"


def csv_string(rows):
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerows(rows)
    return output.getvalue()


def run():
    init_db()
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer((host, port), FinanceHandler)
    print(f"Sistema financeiro rodando em http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    run()

